import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_inv_per_supplier = {}
product_under_10_inv = {}

for product_row in range(2,product_list.max_row + 1):
    inventory = product_list.cell(product_row, 2).value
    supplier_name = product_list.cell(product_row, 4).value
    price = product_list.cell(product_row, 3).value
    product_id = product_list.cell(product_row,1).value
    inventory_price = product_list.cell(product_row,5)

    # cantidad de productos por supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        print(f"Adding a new supplier {supplier_name}")
        products_per_supplier[supplier_name] = 1

    # calculo del precio de cada compra por supplier
    if supplier_name in total_inv_per_supplier:
        total_inv_per_supplier[supplier_name] = total_inv_per_supplier[supplier_name] + (inventory * price)
    else:
        print(f"Adding a new supplier {supplier_name}")
        total_inv_per_supplier[supplier_name] = inventory * price

    # products with inv under 10 units
    if inventory < 10:
        product_under_10_inv[int(product_id)] = int(inventory)

    # add value for total inventory for row
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_inv_per_supplier)
print(product_under_10_inv)

inv_file.save("inventory_with_total_value.xlsx")